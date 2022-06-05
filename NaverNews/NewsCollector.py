import sys

import os
from pathlib import Path

import pymysql
from PyQt5.QtWidgets import QApplication, QWidget, QToolTip, QPushButton
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QFont
from PyQt5 import uic
import openpyxl
from openpyxl import Workbook
import pandas as pd
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

import PyQt5
from PyQt5 import QtWidgets
import urllib
import urllib.request as req
import requests
from bs4 import BeautifulSoup
import math

import smtplib

from email import encoders
from email.utils import formataddr
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# sys.path.append('C:/DataDisk/Python/NaverNews/naver_news_collector.py')
# import naver_news_collector

class AppDemo(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi("C:/DataDisk/Python/NaverNews/NewsCollector.ui", self)

        self.btnSearchSubject.clicked.connect(self.searchsubject)
        self.btnSaveSubject.clicked.connect(self.savesubject)
        self.btnSearchNews.clicked.connect(self.searchnews)
        self.btnResultReview.clicked.connect(self.resultreview)
        self.btnResultUpdate.clicked.connect(self.resultupdate)
        self.btnResultUpdate.clicked.connect(self.resultupdate)
        self.btnSendEMail.clicked.connect(self.sendmail)

        QToolTip.setFont(QFont("SansSerif", 10))

        self.btnSaveSubject.setToolTip(
            '"NewsCollector.xlsx"에 기록된 검색어를 Mysql DB wnews "search_subject" 테이블에 저장합니다.'
        )
        self.btnSearchSubject.setToolTip(
            '검색어를 <b>"NewsCollector.xlsx"</b>에 기록합니다. 파일을 열고 검색어를 확인해 주세요'
        )
        self.btnSearchNews.setToolTip('뉴스를 검색하고 Mysql DB wnews "wwindow" 테이블에 저장합니다.')
        self.btnResultReview.setToolTip(
            'DB에 저장된 검색결과를 <b>"NewsCollector.xlsx"</b>에 기록합니다. 파일을 열고 필요한 뉴스인지 확인해 주세요'
        )
        self.btnResultUpdate.setToolTip("뉴스를 관련된 정보인지 검토한 결과를 업데이트")
        self.btnSendEMail.setToolTip("뉴스를 이메일로 송부합니다.")

        # self.ui.statusbar.showMessage("Status: Idle")

        self.setWindowTitle("네이버 뉴스 검색기")
        self.show()

    ######################################################################################
    def sendmail(self):       

        smtp_address = ("smtp.gmail.com", 587)  # SMTP 서버 호스트, 포트
        # SMTP 계정 정보
        username = "vitrolee@gmail.com"
        # Gmail 계정
        password = "yylridasalpqwpmk"
        # Gmail 앱 비밀번호

        # 메일 송/수신자 정보
        from_addr = ("이정로", "vitrolee@gmail.com")  # 보내는 사람 주소. (이름, 이메일)
        # 메일 제목
        subject = "Windows News를 전달드립니다."
        
        sql1 = "<html><head><title>Windows News</title></head><body><!-- 헤더 부분 --><header><center><h1><p style='line-height: 1px;'>Windows News</p></h1><h3><p style='line-height: 1x;'>한국창호협회</p></h3></center><hr align='center' width=20%></hr></header><br><br><!-- 본문 부분 --><section><article><center>"
        sql2 = "<style type='text/css'>.tg  {border-collapse:collapse;border-color:#9ABAD9;border-spacing:0;}.tg td{background-color:#EBF5FF;border-color:#9ABAD9;border-style:solid;border-width:1px;color:#444;  font-family:Arial, sans-serif;font-size:14px;overflow:hidden;padding:1px 20px;word-break:normal;}.tg th{background-color:#409cff;border-color:#9ABAD9;border-style:solid;border-width:1px;color:#fff;  font-family:Arial, sans-serif;font-size:14px;font-weight:normal;overflow:hidden;padding:1px 20px;word-break:normal;}.tg .tg-r8mc{background-color:#D2E4FC;border-color:inherit;color:#00F;text-align:left;text-decoration:underline;vertical-align:top}.tg .tg-c3ow{border-color:inherit;text-align:left;vertical-align:top}.tg .tg-jn94{background-color:#D2E4FC;border-color:inherit;text-align:left;vertical-align:bottom}.tg .tg-za14{border-color:inherit;text-align:left;vertical-align:bottom}.tg .tg-wucc{border-color:inherit;color:#00F;text-align:left;text-decoration:underline;vertical-align:top}</style>"
        sql3 = "<table class='tg'>"
        sql4 = "<thead><tr><th class='tg-c3ow'>날자</th><th class='tg-c3ow'>언론사</th><th class='tg-c3ow'>기사제목</th></tr></thead>"
        sql5 = "<tbody>"
        sql6 = ""  # <tr><td class='tg-jn94'>2021-12-20</td><td class='tg-jn94'><center>헤럴드경제</center></td>    <td class='tg-r8mc'><a href='http://news.heraldcorp.com/view.php?ud=20211220000456'><span style='text-decoration:underline;color:blue'>LX하우시스, 종합 실내리모델링 ‘성큼’</span></a></td></tr>"
        conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="wnews")
        cur = conn.cursor()
        cur.execute("use wnews")
        #############################################################################################################
        #############################################################################################################
        #############################################################################################################
        sql = 'select cnews_date, info_press, naver_link, news_title from wwindows'
        sql+= ' where news_link is not null and related="1" '

        sql+= ' and sent is null'
        sql+= ' order by cnews_date ASC'
        #sql+= ' and cnews_date >= "20220502" order by cnews_date desc'
        #############################################################################################################
        #############################################################################################################
        #############################################################################################################
        # 기간설정이 필요
        cur.execute(sql)
        rs = cur.fetchall()
        result = pd.DataFrame(rs)

        tablebody = ""
        for kk in range(len(result)):
            cnewsdate = result.iat[kk, 0]
            infopress = result.iat[kk, 1]
            naverlink = result.iat[kk, 2]
            newstitle = result.iat[kk, 3]
            tablebody += f"<tr><td class='tg-jn94'>{cnewsdate}</td><td class='tg-jn94'><center>{infopress}</center></td>    <td class='tg-r8mc'><a href='{naverlink}'><span style='text-decoration:underline;color:blue'><left>{newstitle}</left></span></a></td></tr>"
        sql7 = "</tbody></table></center><br><br><hr size='10' align='center' width=55% noshade='noshade'></hr><!-- 사이트 정보 부분 --><footer><center><small>한국창호협회 vitrolee@gmail.com : 협회밴드 <em class='bandAddressText'> https://band.us/@kwindows</small></center> </footer></body></html>"
        
        # 첨부파일
        attachments = [
            # os.path.join( os.getcwd(), 'storage', 'example.py' )
            # os.path.join(os.chdir(r'C:\Users\JRLee\Downloads\source\report', 'storage', 'WWindows.xlsx' ))
        ]

        conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="wnews")

        cur = conn.cursor()
        cur.execute("use wnews")
        #############################################################################################################
        #############################################################################################################
        #############################################################################################################


        if self.txtToWhom.text()=="All" or self.txtToWhom.text()=="ALL":
            sql = "select 이름, 메일주소 from email_address" #  where 분류='beta'""
        else:
            sql = "select 이름, 메일주소 from email_address where 이름='이정로'"
        #############################################################################################################
        #############################################################################################################
        #############################################################################################################
        cur.execute(sql)

        rs = cur.fetchall()
        df_Name_Address = pd.DataFrame(rs)

        to_addrs = ""  # ('이정로', 'vitrolee@naver.com')

        for kk in range(0, len(df_Name_Address)):
            # 메일 내용         
            #############################################################################################################
            #############################################################################################################
            #############################################################################################################
            sql0 ="" #"<!DOCTYPE html></p><h3>이제는 완연한 봄날입니다. 또한 여기 저기 봄꽃 소식이 전해지고 있습니다.</h3></p>"
            #sql0 +=f" 이제 코로나 바이러스도 매일 감소한다는 소식입니다. 이미 고생하신 분들은 봄날을 만끽하며 올해의 계획을 다시 점검하고 계시죠"
            #sql0 +=f"<br> 아직 코로나와 가까이 하지 않으셨다면 조금만 조심하시면 됩니다. 얼마나 크나 큰 행운입니까. 주변에 모두 좋은 분들과 인생을 즐기고 계십니다."
            #sql0 +=f"<br> 존경하는 " + df_Name_Address.iat[kk, 0] + "님에게 행운이 가득한 봄날이 계속 되길 기원합니다.<br>"            
            #############################################################################################################
            #############################################################################################################
            #############################################################################################################
            body = sql0 + sql1 + sql2 + sql3 + sql4 + sql5 + sql6 + tablebody + sql7

            to_addrs = [("존경하는 " + df_Name_Address.iat[kk, 0] + "님", df_Name_Address.iat[kk, 1])]
            self.lblProgress.setText(f"메일을 {to_addrs} 에 발송하였습니다.")
            mail = NewsMail(smtp_address, from_addr, to_addrs, debug=False)
            mail.setAuth(username, password)
            mail.setMail(subject, body, "html", attachments=attachments)
            mail.send()     
            print(to_addrs[0][0] + "(" + to_addrs[0][1] + ") 에게 "  + "Successfully sent the mail!!!" )   

        #update wnews.wwindows set sent='sent'
        if self.txtToWhom.text()=="All" or self.txtToWhom.text()=="ALL":
            dt_now = datetime.datetime.now()
            p_time = dt_now.strftime("%Y%m%d")
            sql = f"update wnews.wwindows set sent='{p_time}' where sent is null"

            cur.execute(sql)
            conn.commit()
            

        print("메일 보내기가 완료되었습니다!!!" )   
        print("메일 보내기가 완료되었습니다!!!" )  
        print("메일 보내기가 완료되었습니다!!!" )  

    def searchsubject(self):
        conn = pymysql.connect(
            host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
        )
        try:
            cur = conn.cursor()
            cur.execute("use wnews")
            sql = "select subject, subsubject from search_subject"
            cur.execute(sql)

            rs = cur.fetchall()
            wb = load_workbook("C:/DataDisk/Python/NaverNews/NewsCollector.xlsx")

            wb.active
            wb.remove(wb["search_subject"])
            wb.create_sheet("search_subject")
            ws = wb["search_subject"]
            ws.append(("subject", "subsubject"))

            ii = 0
            for row in rs:
                ws.append(row)
            wb.save("C:/DataDisk/Python/NaverNews/NewsCollector.xlsx")
            self.lblProgress.setText("검색어를 'C:/DataDisk/Python/NaverNews/NewsCollector.xlsx' 에 보냈습니다.")
        finally:
            conn.close()
            wb.close()
        return

    def savesubject(self):
        conn = pymysql.connect(
            host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
        )
        try:
            data = pd.read_excel(
                "C:/DataDisk/Python/NaverNews/NewsCollector.xlsx",
                sheet_name="search_subject",
            )

            cur = conn.cursor(pymysql.cursors.DictCursor)
            sql = "delete from search_subject"
            cur.execute(sql)

            sql = "insert into search_subject (subject, subsubject) values(%s, %s)"

            for idx in range(len(data)):
                cur.execute(sql, tuple(data.values[idx]))
            self.lblProgress.setText("검색어를 'C:/DataDisk/Python/NaverNews/NewsCollector.xlsx' 에서 'search_subject' 테이블에 업로드하였습니다.")
            conn.commit()

            # 종료
            cur.close()
            conn.close()
        finally:
            pass
        return

    def searchnews(self):
        page_num = 50
        conn = pymysql.connect(
            host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
        )

        cur = conn.cursor()
        cur.execute("use wnews")
        sql = "select subject, subsubject from search_subject"
        cur.execute(sql)
        rs = cur.fetchall()
        result = pd.DataFrame(rs)

        print(result.loc[0][0])
        for kk in range(len(result)):
            subject = result.iat[kk, 0]
            subsubject = result.iat[kk, 1]
            
            news_count = 0
            for page in range(0, page_num):
                if page == 0:
                    url = f"https://search.naver.com/search.naver?where=news&query={subsubject}&sm=tab_opt&sort=1&photo=0&field=0&pd=0&ds=&de=&docid=&related=0&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so%3Add%2Cp%3Aall&is_sug_officeid=0"
                else:
                    url = f"https://search.naver.com/search.naver?where=news&sm=tab_pge&query={subsubject}&sort=1&photo=0&field=0&pd=0&ds=&de=&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:dd,p:all,a:all&start={page*10+1}"

                header = {"User-agent": "Mozila/2.0"}
                response = requests.get(url, headers=header)
                html = response.text
                soup = BeautifulSoup(html, "html.parser")
                titles = soup.select(".news_wrap")

                for title in titles:
                    news_time = ""
                    info_press = ""
                    news_title = ""
                    news_link = ""
                    naver_link = ""
                    try:
                        info_press = title.find(
                            "a", attrs={"class": "info press"}
                        ).text.strip()
                    except:
                        pass
                    news_title = title.select_one(".news_tit").text.strip()
                    news_link = title.select_one(".news_tit")["href"].strip()
                    infogroups = title.select(".info_group")

                    for infogroup in infogroups:
                        for ii in range(0, len(infogroup) - 2):
                            if (
                                " 전" in infogroup.select(".info")[ii].text
                                or "." in infogroup.select(".info")[ii].text
                            ):
                                news_time = infogroup.select(".info")[ii].text
                            elif "네이버뉴스" in infogroup.select(".info")[ii].text:
                                naver_news = infogroup.select(".info")[ii].text
                                naver_link = infogroup.select(".info")[ii].attrs["href"]
                            else:
                                continue

                    cnews_date = calc_released_date(news_time)
                    # 문자 ", ' 문자를 제거하는 과정 한개가 아니고 여러개가 있을 수 있으므로
                    for ii in range(1, 10):
                        news_title = news_title.replace('"', "")
                        cnews_date = cnews_date.replace('"', "")
                        info_press = info_press.replace('"', "")
                        news_link = news_link.replace('"', "")

                        news_title = news_title.replace("'", "")
                        cnews_date = cnews_date.replace("'", "")
                        info_press = info_press.replace("'", "")
                        news_link = news_link.replace("'", "")

                        news_title = news_title.replace("‘", "")
                        cnews_date = cnews_date.replace("‘", "")
                        info_press = info_press.replace("‘", "")
                        news_link = news_link.replace("‘", "")

                    news_count += 1
                    print(news_count, " ", news_title)
                    print(subsubject, "  ", cnews_date, "  ", info_press, news_link)

                    print("-" * 100)

                    news_content = ""
                    dt_now = datetime.datetime.now()
                    p_time = dt_now.strftime("%Y-%m-%d %H:%M:%S")

                    # 얻어진 뉴스정보를 저장
                    conn = pymysql.connect(
                        host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
                    )
                    cur = conn.cursor()
                    cur.execute("use wnews")

                    sql = "INSERT ignore INTO wwindows"
                    sql += " (news_date, cnews_date, info_press, news_title, news_link, news_content, naver_link, subject, subsubject, p_time)"
                    sql += " values ("
                    sql += " '" + str(news_time) + "'"
                    sql += ", '" + str(cnews_date) + "'"
                    sql += ", '" + info_press + "'"
                    sql += ", '" + news_title + "'"
                    sql += ", '" + news_link + "'"
                    sql += ", '" + news_content + "'"
                    sql += ", '" + naver_link + "'"
                    sql += ", '" + subject + "'"
                    sql += ", '" + subsubject + "'"
                    sql += ", '" + str(p_time) + "');"

                    cur.execute(sql)
                    conn.commit()
                    cur.close()
                    conn.close()
        self.lblProgress.setText(f"주검색어: {subject}, 부검색어: {subsubject} 를 검색을 완료하였습니다.")
        self.resultreview()
        
    def resultreview(self):
        self.lblProgress.setText("resultreview 에서 시작합니다.")
        conn = pymysql.connect(
            host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
        )

        cur = conn.cursor()
        cur.execute("use wnews")

        # 네이버에 뉴스가 없는 경우 언론사의 링크를 네이버 링크에 연결'

        sql = 'update wwindows set naver_link = news_link where (naver_link ="") or (naver_link is null)'
        cur.execute(sql)
        conn.commit()

        sql = 'update wwindows T INNER JOIN wwindows S on T.news_link=S.news_link set T.cnews_link =  concat("=HYPERLINK(""", S.news_link, """, """, S.news_title, """) ") where T.cnews_link is null'
        cur.execute(sql)
        conn.commit()

        try:
            # conn=pymysql.connect(host='127.0.0.1',user='root',passwd='JRLEE',db='wnews')
            #cur = conn.cursor()
            #cur.execute("use wnews")

            sql = "select cnews_date, info_press, subject, subsubject, related, news_title, cnews_link, news_link from wwindows where naver_link <>'' and (related is null or related='1') order by cnews_date desc"
            cur.execute(sql)
            rs = cur.fetchall()

            # wb = Workbook()
            wb = load_workbook("C:/DataDisk/Python/NaverNews/NewsCollector.xlsx")
            wb.active
            wb.remove(wb["result_reviw"])
            wb.create_sheet("result_reviw")
            ws = wb["result_reviw"]

            # 첫행 입력
            ws.append(
                (
                    "cnews_date",
                    "info_press",
                    "subject",
                    "subsubject",
                    "related",
                    "news_title",
                    "cnews_link",
                    "news_link",                    
                )
            )

            # DB 모든 데이터 엑셀로
            ii = 0
            for row in rs:
                ws.append(row)

            wb.save("C:/DataDisk/Python/NaverNews/NewsCollector.xlsx")
            self.lblProgress.setText('검색된 뉴스를 "C:/DataDisk/Python/NaverNews/NewsCollector.xlsx" 저장하였습니다.')

            absolutePath = Path("C:/DataDisk/Python/NaverNews/NewsCollector.xlsx").resolve()
            os.system(f'start excel.exe "{absolutePath}"')
            print("검색이 완료되었습니다.")

        finally:
            conn.close()
            wb.close()
        return
        print("NewsCollector.xlsx 이름으로 저장되었습니다.")
    def resultupdate(self):
        conn = pymysql.connect(
            host="127.0.0.1", user="root", passwd="JRLEE", db="wnews"
        )
        try:
            data = pd.read_excel(
                "C:/DataDisk/Python/NaverNews/NewsCollector.xlsx",
                sheet_name="result_reviw",
            )

            cur = conn.cursor(pymysql.cursors.DictCursor)
            # sql = "update wwindows set related = %s where news_link=%s"

            for idx in range(len(data)):
                if math.isnan(data.iat[idx, 4]) == False:
                    sql = f'update wwindows set related = "{str(int(data.iat[idx,4]))}" where news_link="{data.iat[idx,7]}"'
                    cur.execute(sql)
                    conn.commit()
                else:
                    continue
            conn.commit()
            cur.close()
            conn.close()
            self.lblProgress.setText('검토된 뉴스를 "C:/DataDisk/Python/NaverNews/NewsCollector.xlsx" 에서 데이터베이스 "wnews", "wwindows" 테이블에 저장하였습니다.')
        finally:
            pass
        return
        print("업데이트 되었습니다.")
    # def sendemail(self):
    # continue
class NewsMail(object):
    ENCODING = "UTF-8"

    host = None
    port = None
    from_addr = None
    to_addrs = None
    message = None

    debug = False
    _isLogin = False
    _existAttachments = False

    def __init__(self, address: tuple, from_addr: str, to_addrs: list, debug=False, **kwargs):
        self.host, self.port = address
        self.from_addr = from_addr
        self.to_addrs = to_addrs

        # 인증 계정
        if kwargs.get("vitrolee") is not None:
            self.setAuth(kwargs.get("vitrolee"), kwargs.get("yylridasalpqwpmk"))

        # 디버그 모드
        self.debug = debug is not False

        # 첨부파일로 전송 불가능한 확장자
        self.blocked_extensions = (".ade",".adp",".apk",".appx",".appxbundle",".bat",".cab",".chm",".cmd",".com",".cpl",
            ".dll",".dmg",".exe",".hta",".ins",".isp",".iso",".jar",".js",".jse",".lib",".lnk",".mde",".msc",".msi",
            ".msix",".msixbundle",".msp",".mst",".nsh",".pif",".ps1",".scr",".sct",".shb",".sys",".vb",".vbe",".vbs",
            ".vxd",".wsc",".wsf",".wsh",)

    def setAuth(self, username, password):
        self._isLogin = True
        self.username = "vitrolee"  # username
        self.password = "yylridasalpqwpmk"  # password

    def setMail(self, subject, body, body_type="plain", attachments=None):
        """
        Content-Type:
            - multipart/alternative: 평문 텍스트와 HTML과 같이 다른 포맷을 함께 보낸 메시지
            - multipart/mixed: 첨부파일이 포함된 메시지

        REF:
            - https://ko.wikipedia.org/wiki/MIME#Content-Type
        """
        # 첨부파일 여부
        self._existAttachments = attachments is not None
        self.content_sub_type = "mixed" if self._existAttachments else "alternative"

        # 메일 콘텐츠 설정
        self.message = MIMEMultipart(self.content_sub_type)

        # 받는 사람 주소 설정. [( 이름 <이메일> )]
        self.FROM_ADDR_FMT = formataddr(self.from_addr)
        self.TO_ADDRS_FMT = ",".join([formataddr(addr) for addr in self.to_addrs])


        # 메일 헤더 설정
        self.message.set_charset(self.ENCODING)
        self.message["From"] = self.FROM_ADDR_FMT
        self.message["To"] = self.TO_ADDRS_FMT
        self.message["Subject"] = subject

        # 메일 콘텐츠 - 내용
        self.message.attach(MIMEText(body, body_type, self.ENCODING))

        # 메일 콘텐츠 - 첨부파일
        if self._existAttachments:
            self._attach_files(attachments)

        return self

    def _attach_files(self, attachments):
        """
        Content-disposition:
            - 파일명 지정
        MIME type:
            - application/octect-stream: Binary Data
        REF:
            - https://www.freeformatter.com/mime-types-list.html
        """

        for attachment in attachments:
            attach_binary = MIMEBase("application", "octect-stream")
            try:
                binary = open(attachment, "rb").read()  # read file to bytes

                attach_binary.set_payload(binary)
                encoders.encode_base64(
                    attach_binary
                )  # Content-Transfer-Encoding: base64

                #filename = os.path.basename(attachment)
                filename = sys.path.basename(attachment)
                
                attach_binary.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename=(self.ENCODING, "", filename),
                )

                self.message.attach(attach_binary)
            except Exception as e:
                print(e)

    def send(self):
        session = None
        
        try:
            # 메일 세션 생성
            session = smtplib.SMTP(self.host, self.port)
            session.set_debuglevel(self.debug)
            session.ehlo()

            # SMTP 서버 계정 인증
            if self._isLogin:
                session.starttls()
                session.login(self.username, self.password)

            # 메일 발송
            session.sendmail(
                self.FROM_ADDR_FMT, self.TO_ADDRS_FMT, self.message.as_string()
            )            

        except Exception as e:
            print(e)
        finally:
            if session is not None:
                session.quit()

def calc_released_date(calc_time):
    now = datetime.datetime.now()
    if "분 전" in calc_time:
        cnews_date = now - relativedelta(minutes=int(calc_time.replace("분 전", "")))
    elif "시간 전" in calc_time:
        cnews_date = now - relativedelta(hours=int(calc_time.replace("시간 전", "")))
    elif "일 전" in calc_time:
        cnews_date = now - relativedelta(days=int(calc_time.replace("일 전", "")))
    elif "." in calc_time:
        format = "%Y.%m.%d."
        cnews_date = datetime.datetime.strptime(calc_time, format)
    else:
        cnews_date = now

    cnews_date = cnews_date.strftime("%Y%m%d")
    return cnews_date


def naver_news_contents():
    # 뉴스 상세내용을 가져오는 과정으로 현재 사용하지 않고 있음
    conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="wnews")
    cur = conn.cursor()
    cur.execute("use wnews")
    sql = "select naver_link,  char_length(naver_link) from wwindows where char_length(naver_link) > 10 and char_length(news_content)<10;"
    cur.execute(sql)

    res = cur.fetchall()

    conn.commit()
    cur.close()
    conn.close()

    for data in res:
        print(cur.fetchone())
        # print(data)
        url = data[0]
        html = urllib.request.urlopen(url).read()
        soup = BeautifulSoup(html)

        # kill all script and style elements
        for script in soup(["script", "style"]):
            script.extract()  # rip it out

        # get text
        text = soup.get_text()

        # break into lines and remove leading and trailing space on each
        lines = (line.strip() for line in text.splitlines())
        # break multi-headlines into a line each
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        # drop blank lines
        text = "\n".join(chunk for chunk in chunks if chunk)

        print(text)
        
        header = {"User-agent": "Mozila/2.0"}

        response = requests.get(url, headers=header)
        html = response.text
        soup = BeautifulSoup(html, "html.parser")

        # articeBody
        news_content = soup.find("div", attrs={"id": "articeBody"}).get_text(
            " "
        )  # text

        # for ii in range(1,10):
        #     news_content = news_content.replace('"',"")
        #     news_content = news_content.replace("'","")
        #     news_content = news_content.replace("/n","")

        #     text_parts = soup.findAll(text=True)

        # news_content=soup.select('.news_wrap')
        news_content = soup.find("div", attrs={"id": "articeBody"})
        script_tag = news_content.find_all(
            ["script", "style", "header", "footer", "form"]
        )
        for script in script_tag:
            script.extract()
        content = soup.get_text("/n", strip=True)
        print(content)

        script_tag = soup.find_all(["script", "style", "header", "footer", "form"])
        for script in script_tag:
            script.extract()
        content = soup.get_text("/n", strip=True)
        print(content)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    demo = AppDemo()
    demo.show()

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("Closing")

