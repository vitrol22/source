import requests
from bs4 import BeautifulSoup
import pymysql
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def get_soup(url):
    header = {
        "User-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36"
    }
    res = requests.get(url, headers=header)
    if res.status_code == 200:
        return BeautifulSoup(res.text, "html.parser")



def mainInfo(s):
    conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="scholar")

    elements = s.select(".ui_listing_info")

    ii = 0
    for element in elements:
        title = "Empty"
        title_link = "Empty"
        publish_year ="0"
        author = "Empty"
        jounal = "Empty"
        referenced = "0"

        title = element.select(".ui_listing_subtit")[0].text
        title_link = (
            "https://m.academic.naver.com"
            + element.select(".ui_listing_subtit")[0]["href"]
        )
        try:
            sub_elements = element.select(".ui_listing_desc")[0]
            # for sub_element in sub_elements:
            #    print(sub_element)

            if len(sub_elements)==9:
                try:
                    author = sub_elements("span")[1].text
                except:
                    author ="Empty"
                try:
                    publish_year = element.select("div.ui_listing_info > div > span")[0].text                 
                    if len(publish_year)==4:
                        publish_year=publish_year
                    else:
                        publish_year=''
                except:
                    publish_year =""
                try:
                    jounal=sub_elements.select(".ui_listing_source")[2].text
                    jounal=jounal.replace("\n","")
                except:
                    jounal ="Empty"
                try:
                    
                    referenced = element.select(".ui_listing_cited")[0].text
                    referenced = referenced.replace("회 피인용", "")
                except:
                    referenced ="0"
            else:
                pass
        except:
            pass #continue


        #print(ii, "번째 문서", "new article! __________________________________________________________________________", )
        try:
            title = title.replace('"', "")
            title = title.replace("'", "")
            title_link = title_link.replace('"', "")
            title_link = title_link.replace("'", "")
            publish_year = publish_year.replace('"', "")
            publish_year = publish_year.replace("'", "")
            author = author.replace('"', "")
            author = author.replace("'", "")
            jounal = jounal.replace('"', "")
            jounal = jounal.replace("'", "")
            referenced = referenced.replace('"', "")
            referenced = referenced.replace("'", "")
            if referenced=='':
                referenced = 0
        except:
            print("정보에서 발견하지 못한 항목이 있습니다.")
        # print(len(sub_elements))
        # print("title => ", title)
        # print("title link => ", title_link)
        # print("publish year => ", publish_year)
        # print("author => ", author)
        # print("jounal => ", jounal)
        # print("referenced => ", referenced)
        ii += 1

        try:
            save_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur = conn.cursor(pymysql.cursors.DictCursor)

            sql = "INSERT ignore INTO naver"
            #sql += " (idnaver, subject, title, title_link, publish_year, author, jounal, referenced, save_time)"
            sql += " (subject, title, title_link, publish_year, author, jounal, referenced, save_time)"
            sql += " values ("
            #sql += " '" + str(ii) + "'"            
            sql += " '" + subject + "'"
            sql += ", '" + title + "'"
            sql += ", '" + title_link + "'"
            sql += ", '" + publish_year[:10] + "'"
            sql += ", '" + author + "'"
            sql += ", '" + jounal + "'"
            # sql += ", '" + abstract + "'"
            sql += ", '" + referenced + "'"
            sql += ", '" + save_time + "'"            
            sql += ");"
            # print(sql)
            #print(title)
            cur.execute(sql)
            conn.commit()
        except pymysql.err.InternalError as e:
            print("Error code:", e.errno)  # error number
            print("SQLSTATE value:", e.sqlstate)  # SQLSTATE value
            print("Error message:", e.msg)  # error message
            print("Error:", e)  # errno, sqlstate, msg values
            s = str(e)
            print("Error:", s)  # errno, sqlstate, msg values

        finally:
            # print(title)
            cur.close()
            continue


def detail_search():
    cur = conn.cursor()
    cur.execute("use scholar")
    sql = (
        f"select title_link from scholar.naver where subject='{subject}' and (abstract is null or abstract ='Empty') order by idnaver;"
    )
    cur.execute(sql)

    rs = cur.fetchall()
    result = pd.DataFrame(rs)
    url = ""

    for kk in range(len(result)):
        total_number=len(result)
        url = [result.iat[kk, 0]][0]
        s = get_soup(url)

        author = "Empty"
        doi = "Empty"
        categoryT="Empty"
        keywords="Empty"
        abstract = "Empty"

        try:
            if s.select(".ui_listdetail.type2 dl")[0]("dt")[0].text == "저자":
                author = s.select(".ui_listdetail.type2 dl")[0]("dd")[0].text
                author = author.replace('"', "")
                author = author.replace("'", "")
        except:
            author = "Empty"
            #print("author 정보가 없어요", url)

        try:
            for mm in range(0, len(s.select(".ui_listdetail.type2 dl")[0]("dt"))):
                if s.select(".ui_listdetail.type2 dl")[0]("dt")[mm].text == "저자":
                    author = s.select(".ui_listdetail.type2 dl")[0]("dd")[mm].text
                    author = author.replace('"', "")
                    author = author.replace("'", "")
                elif s.select(".ui_listdetail.type2 dl")[0]("dt")[mm].text == "DOI":
                    doi = s.select(".ui_listdetail.type2 dl")[0]("dd")[mm].text
                    doi = doi.replace('"', "")
                    doi = doi.replace("'", "")
        except:
            author = "Empty"
            #print("doi 정보가 없어요", url)
            pass#continue

        try:            
            for mm in range(0, len(s.select(".ui_listdetail.type2 dl")[2]("dt"))):
                if s.select(".ui_listdetail.type2 dl")[2]("dt")[mm].text == "주제분야":
                    categoryT = s.select(".ui_listdetail.type2 dl")[2]("dd")[mm].text
                    #categoryT = categoryT.replace('\n> \n', ">")
                    for mm in range(1,10):
                        categoryT = categoryT.replace('\r\n', '')
                        categoryT = categoryT.replace("\n", "")    
                        categoryT = categoryT.replace("\r", "")
                        categoryT = categoryT.replace("\text", "")    
                        categoryT = categoryT.replace("\t", "")

                elif s.select(".ui_listdetail.type2 dl")[2]("dt")[mm].text == "키워드":
                    keywords = s.select(".ui_listdetail.type2 dl")[2]("dd")[mm].text
                    for mm in range(1,10):
                        keywords = keywords.replace('\r\n', '')
                        keywords = keywords.replace('\r', ' ')
                        keywords = keywords.replace('\n', ' ')
                        keywords = keywords.replace('&nbsp;', " ")      # 1개의 Tab을 지우는 과정
                        keywords = keywords.replace("&ensp", "  ")      # 2개의 Tab을 지우는 과정
                        keywords = keywords.replace('&emsp', "   ")     # 3개의 Tab을 지우는 과정
                        keywords = keywords.replace('<br>', " ")     # 3개의 Tab을 지우는 과정
                        keywords = keywords.replace('&#09', " ")
                        keywords = keywords.replace('<pre>', " ")
                        keywords = keywords.replace('<p>', " ")
                        keywords = keywords.replace('text-indent', " ")
                        keywords = keywords.replace('\text', " ")
                        #keywords = keywords.replace('\t', " ")
                        keywords = keywords.replace('"', "")
                        keywords = keywords.replace("'", "")
        except:
            #print("categoryT 정보가 없어요", url)
            #print()
            pass
            # continue

        try:            
            abstract = s.select("#div_abstract")[0]("p")[0].text
            for mm in range(1,10):
                abstract = abstract.replace('\r\n', ' ')
                abstract = abstract.replace('\n', ' ')
                abstract = abstract.replace('\r', ' ')
                abstract = abstract.replace('"', "")
                abstract = abstract.replace("'", "")
                abstract = abstract.replace('&nbsp;', " ")      # 1개의 Tab을 지우는 과정
                abstract = abstract.replace("&ensp", "  ")      # 2개의 Tab을 지우는 과정
                abstract = abstract.replace('&emsp', "   ")     # 3개의 Tab을 지우는 과정
                abstract = abstract.replace('<br>', " ")     # 3개의 Tab을 지우는 과정                
                abstract = abstract.replace("\text", "").strip()
             
            if len(abstract)<1:
                abstract='Empty'
            else:
                abstract=abstract
                #continue # abstract = abstract
        except:
            
            abstract = "Empty"
            #print("abstract 정보가 없어요", url)

        try:
            # author="Empty" if author="" else author=author
            cur = conn.cursor(pymysql.cursors.DictCursor)
            sql = "UPDATE naver set"
            sql += " author= '" + author[:100] + "'"
            sql += ", DOI= '" + doi + "'"
            sql += ", category= '" + categoryT[:100]  + "'"
            sql += ", keywords= '" + keywords[:100] + "'"
            sql += ", abstract= '" + abstract + "'" 
            sql += " where title_link = '" + url + "'"
            sql += ";"

            try:
                cur.execute(sql)
            except:
                # pymysql.err.InternalError as e:
                print("저장실패")
                #code, msg = e.args

            conn.commit()
            print("Detail Search", 'suject=', subject, " ", kk,"/",total_number, "정상 저장", "  DOI=",doi)

        except:
            print("저장에 실패헸어요", url) 
            #print
        finally:
            cur.close()

def mainStarter(url):
    s = get_soup(url)
    if s == None:
        docCount =0
    else:
        docCount = s.select("li.ui_tabnavi_item.on > a > span")[0].text
        docCount = int(docCount.replace(",", ""))

    sort = "0"
    doctype = ""
    year = ""
    page = "0"

    if docCount == 0:
        pass
    elif docCount < 2000:
        for page in range(0, int(docCount / 10) + 1):
            #print("page=",page,"")
            url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={sort}&searchType=1&refineType=exist&docType=&thesisLv=&journalLv=&access=&year=&category=&journal=&source=&page={page}"
            s = get_soup(url)
            if s == None:
                continue
            else:
                #print("year=",year,"",'sort=', sort, "page=",page,"")
                print(subject, "Main Search < 2000", docCount,"/", docCount, "page=",page,"")
                mainInfo(s)
    elif docCount <= 4000:
        cdocCount =2000 if docCount >= 2000 else docCount    #200 page 까지만 접근가능하므로
        for sort in range(3, 5):
            for page in range(0, int(cdocCount / 10) + 1):            
                #print("year=",year,"","page=",page,"")
                url = f"https://academic.naver.com/search.naver?query={subject}&field=0&sort={sort}&searchType=1&refineType=exist&docType=&thesisLv=&journalLv=&access=&year=&category=&journal=&source=&page={page}"
                #print(url)
                s = get_soup(url)
                if s == None:
                    continue
                else:
                    #print("year=",year,"",'sort=', sort, "page=",page,"")
                    print(subject, "Main Search < 4000",'sort=', sort, cdocCount,"/", docCount, "page=",page,"")
                    mainInfo(s)

    elif docCount > 4000:
        #for year in range(2018, 2023):
        for year in range(1900, 2023):    
            year = str(year)
            url = f"https://academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
            s = get_soup(url)
            if s == None:
                continue
            else:
                yeardocCount = s.select("li.ui_tabnavi_item.on > a > span")[0].text
                yeardocCount = int(yeardocCount.replace(",", ""))
                cyeardocCount =2000 if yeardocCount >= 2000 else yeardocCount #200 page 까지만 접근가능하므로
                if yeardocCount == 0:
                    pass
                elif yeardocCount<=2000:
                    #yeardocCount =2000 if yeardocCount >= 2000 else yeardocCount
                    for page in range(0, int(cyeardocCount / 10) + 1): 
                        url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
                        s = get_soup(url)
                        if s == None:
                            continue
                        else:
                            print(subject, "Main Search>4000<=2000","year=",year,"", yeardocCount,"/", docCount, "page=",page,"")
                            mainInfo(s)                       
                elif yeardocCount<=4000:
                    for sort in range(3, 5):
                        url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
                        s = get_soup(url)
                        if s == None:
                            continue
                        else:
                            sortyeardocCount = s.select("li.ui_tabnavi_item.on > a > span")[0].text
                            sortyeardocCount = int(sortyeardocCount.replace(",", ""))                            
                            csortyeardocCount =2000 if sortyeardocCount >= 2000 else sortyeardocCount #200 page 까지만 접근가능하므로
                            for page in range(0, int(csortyeardocCount / 10) + 1): 
                                url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
                                s = get_soup(url)
                                print(subject, "Main Search>4000<4000","year=",year,"",'sort=', sort, sortyeardocCount,"/",docCount, "page=",page,"")
                                mainInfo(s)   
                else:
                    for sort in range(0, 5):
                        url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
                        s = get_soup(url)
                        if s == None:
                            continue
                        else:
                            sortyeardocCount = s.select("li.ui_tabnavi_item.on > a > span")[0].text
                            sortyeardocCount = int(sortyeardocCount.replace(",", ""))                            
                            csortyeardocCount =2000 if sortyeardocCount >= 2000 else sortyeardocCount #200 page 까지만 접근가능하므로
                            for page in range(0, int(csortyeardocCount / 10) + 1): 
                                url = f"https://m.academic.naver.com/search.naver?query={subject}&field=0&sort={str(sort)}&searchType=1&refineType=exist&docType={doctype}&thesisLv=&journalLv=&access=&year={year}%3A2000&category=&journal=&source=&page={page}"
                                s = get_soup(url)
                                print(subject, "Main Search>4000>4000","year=",year,"",'sort=', sort, sortyeardocCount,"/", docCount, "page=",page,"")
                                mainInfo(s) 

def resultDump(subject):
    conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="scholar")

    #cur = conn.cursor()
    #cur.execute("use scholar")

    try:
        cur = conn.cursor()
        cur2=conn.cursor()
        cur.execute("use scholar")
        cur2.execute("use scholar")
        #sql = f"SELECT (@cnt := @cnt + 1) AS idnaver, subject, title, title_link, publish_year, author, jounal, referenced, save_time, abstract, DOI, category, keywords, 중요, 메모 from naver where subject='{subject}' order by referenced desc"
        sql = f"SELECT ROW_NUMBER() OVER (ORDER BY length(referenced) DESC, referenced desc) AS idnaver, subject, title, title_link, publish_year, author, jounal, referenced, save_time, abstract, DOI, category, keywords, 중요, 메모 from naver where subject='{subject}' order by length(referenced) DESC, referenced desc"
        cur.execute(sql)
        rs = cur.fetchall()

    # Create Excel (.xlsx) file -----------------------------------------------
        wb = Workbook()
        ws = wb.create_sheet(0)
        ws.title = subject

        ws.append((
                "idnaver", "subject", "title", "title_link", "publish_year", "author", "jounal", "referenced", "save_time", "abstract", "DOI",	"category", "keywords", "중요", "메모"		
            ))
        ii=0
        for row in rs:
            ii+=1
            try:
                ws.append(row)
            except:
                sql2="delete from naver where idnaver=" + str(row[0])
                cur.execute(sql2)
                rs2 = cur2.fetchall()


        workbook_name = subject
        wb.save(workbook_name + ".xlsx")

    finally:
        conn.close()
        wb.close()
    return
    
# mysql workbench 를 가동하여 batchmission을 실행한다.
# delete from batchmission;

#insert batchmission values ("block+copolymer+encapsulation", "main", "5","");
#insert batchmission values ("block+copolymer+encapsulation", "detail", "6","");

if __name__ == "__main__":
    # mysql에 찿을 주제어를 넣고 시작
    #처음은 main
    # 두번째는 Detail 을 찿는다.
    
    conn = pymysql.connect(host="127.0.0.1", user="root", passwd="JRLEE", db="scholar")

    cur = conn.cursor(pymysql.cursors.DictCursor)
    cur.execute("use scholar")

    sql="select * from batchmission where process =''"
    cur.execute(sql)
    rs = cur.fetchall()
    result = pd.DataFrame(rs)
    #conn.commit()
    #cur.close()

    for kk in range(len(result)):
        subject = result.iat[kk, 0]
        mainOrdetail = result.iat[kk, 1]

        
        if mainOrdetail=='main':
            url = f"https://academic.naver.com/search.naver?query={subject}&searchType=1&field=0&docType="
            mainStarter(url)

            dt_now = datetime.datetime.now()
            p_time = dt_now.strftime("%Y%m%d %H:%M:%S")

            sql = "UPDATE batchmission set"
            sql += f" process = '{p_time}'"
            sql += f" where subject = '{subject}'"
            sql += f" and mainOrdetail = '{mainOrdetail}'"
            sql += ";"

            cur.execute(sql)
            conn.commit()

        elif mainOrdetail=='detail':
            detail_search()

            dt_now = datetime.datetime.now()
            p_time = dt_now.strftime("%Y%m%d %H:%M:%S")

            sql = "UPDATE batchmission set"
            sql += f" process = '{p_time}'"
            sql += f" where subject = '{subject}'"
            sql += f" and mainOrdetail = '{mainOrdetail}'"
            sql += ";"

            cur.execute(sql)
            conn.commit()

sql="select * from batchmission"
cur.execute(sql)
rs = cur.fetchall()
result = pd.DataFrame(rs)

for kk in range(len(result)):
    subject = result.iat[kk, 0] 

    resultDump(subject)
    print(subject+"파일에 데이터를 저장하였습니다.")

print("종료하였습니다.")

    # CREATE TABLE mecaro LIKE naver; 
    # INSERT mecaro SELECT * FROM naver;
